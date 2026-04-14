import openpyxl
vf = 0
vp = 5000
taxa = 0.8
periodo = 5

Caminho_Arquivo="Investimentos Unicos.xlsx"
try:
    workbook = openpyxl.load_workbook(Caminho_Arquivo) #Carrega a pasta de trabalho do excel
    print(f"Arquivo {Caminho_Arquivo} carregado com sucesso!")
except FileNotFoundError:
    print (f"Erro: O arquivo {Caminho_Arquivo} não foi encontrado!")
except Exception as e:   
    print(f"Erro: {e}")

planilha_nome = "Compras"
workbook = openpyxl.load_workbook(Caminho_Arquivo)
if planilha_nome in workbook.sheetnames:
    planilha = workbook[planilha_nome]
    vf = vp * (1+taxa)**periodo
    print (f"--- Escrevendo dados na planilha {planilha.title} ---")
    celula_a1=planilha["A5"] #Modificando uma célula existente (Ex: A1)
    valor_antigo_a1 =celula_a1.value
    novo_valor_a1= "Valor Presente"
    celula_a1.value = novo_valor_a1
    print(f"Celula B5 modificada de {valor_antigo_a1} para {novo_valor_a1}")
    celula_a1=planilha["B5"] #Modificando uma célula existente (Ex: A1)
    valor_antigo_a1 =celula_a1.value
    novo_valor_a1= vf
    celula_a1.value = novo_valor_a1
    print(f"Celula B5 modificada de {valor_antigo_a1} para {novo_valor_a1}")
    workbook.save(Caminho_Arquivo)
    print("Alterações salvas com sucesso!")
else:
    print(f"A planilha '{planilha_nome}' não foi encontrada no arquivo.")

