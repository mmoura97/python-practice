from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Projeto"
ws["A1"] = "Periodo"
ws["B1"] = "Fluxo de Caixa"

fluxos = [-5000, 1500, 2000, 2500, 2000]
for i, valor in enumerate(fluxos):
   ws.cell(row=i+2, column=1, value=i)       # período
   ws.cell(row=i+2, column=2, value=valor)   # fluxo
wb.save("Projeto.xlsx")

wb = load_workbook("Projeto.xlsx")
ws = wb.active
fluxos_lidos = []
for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
   fluxos_lidos.append(row[1])

taxa = 0.10  # 10%
vpl = 0
for t, fc in enumerate(fluxos_lidos):
   vpl += fc / (1 + taxa) ** t

ws["A8"] = "VPL"
ws["B8"] = vpl
wb.save("Projeto.xlsx")

print(f"VPL calculado: {vpl:.2f}")
if vpl > 0:
   print("Projeto viável")
else:
   print("Projeto inviável")

